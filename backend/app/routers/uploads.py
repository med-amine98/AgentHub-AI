"""
uploads.py -- File upload router for AgentHub AI.

Supported formats:
  - CSV  -> parsed to JSON list of rows
  - XLSX / XLS -> parsed to JSON list of rows (openpyxl or csv fallback)
  - PDF  -> text extracted via pdfplumber (or basic fallback)
  - TXT / MD / JSON / PY / JS -> read as plain text
  - Images (PNG, JPG, WEBP) -> described by Gemini Vision if key present
"""

import os
import json
import csv
import io
import base64
import uuid
import datetime
from pathlib import Path
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from sqlalchemy.orm import Session

from ..database import get_db
from .. import models, schemas, auth

router = APIRouter(
    prefix="/api/uploads",
    tags=["File Uploads"]
)

UPLOADS_DIR = Path(os.environ.get("UPLOADS_DIR", "./uploads"))
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

MAX_FILE_SIZE = 20 * 1024 * 1024  # 20 MB


def parse_file_content(file_bytes: bytes, mime_type: str, filename: str) -> str:
    ext = Path(filename).suffix.lower()

    # CSV
    if mime_type == "text/csv" or ext == ".csv":
        try:
            text = file_bytes.decode("utf-8", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
            if rows:
                return json.dumps(rows[:500], ensure_ascii=False)
            return text[:10000]
        except Exception:
            return file_bytes.decode("utf-8", errors="replace")[:10000]

    # Excel
    if ext in (".xlsx", ".xls") or "spreadsheet" in mime_type or "ms-excel" in mime_type:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            result = {}
            for sheet_name in wb.sheetnames[:3]:
                ws = wb[sheet_name]
                headers = None
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        headers = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
                    else:
                        if any(c is not None for c in row):
                            rows.append(dict(zip(headers, [str(c) if c is not None else "" for c in row])))
                    if i > 500:
                        break
                result[sheet_name] = rows
            return json.dumps(result, ensure_ascii=False)
        except ImportError:
            try:
                return file_bytes.decode("utf-8", errors="replace")[:10000]
            except Exception:
                return "[Excel file -- install openpyxl for parsing]"
        except Exception as e:
            return f"[Excel parse error: {e}]"

    # PDF
    if mime_type == "application/pdf" or ext == ".pdf":
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                text_parts = []
                for page in pdf.pages[:20]:
                    t = page.extract_text()
                    if t:
                        text_parts.append(t)
            return "\n\n".join(text_parts)[:30000]
        except ImportError:
            try:
                return file_bytes.decode("latin-1", errors="replace")[:10000]
            except Exception:
                return "[PDF file -- install pdfplumber for text extraction]"
        except Exception as e:
            return f"[PDF parse error: {e}]"

    # Images
    if mime_type in ("image/png", "image/jpeg", "image/webp", "image/gif") or ext in (".png", ".jpg", ".jpeg", ".webp"):
        try:
            from google import genai
            api_key = os.environ.get("GEMINI_API_KEY")
            if api_key:
                client = genai.Client(api_key=api_key)
                b64 = base64.b64encode(file_bytes).decode()
                response = client.models.generate_content(
                    model="gemini-2.0-flash",
                    contents=[
                        {"role": "user", "parts": [
                            {"text": "Decris precisement le contenu de cette image. Si elle contient des donnees, un tableau ou un graphique, extrais toutes les valeurs numeriques et categories visibles. Reponds en francais."},
                            {"inline_data": {"mime_type": mime_type, "data": b64}}
                        ]}
                    ]
                )
                return f"[Image decrite par IA] : {response.text}"
        except Exception:
            pass
        return f"[Image binaire -- {len(file_bytes)} octets, type: {mime_type}]"

    # JSON
    if mime_type == "application/json" or ext == ".json":
        try:
            data = json.loads(file_bytes.decode("utf-8", errors="replace"))
            return json.dumps(data, ensure_ascii=False)[:30000]
        except Exception:
            return file_bytes.decode("utf-8", errors="replace")[:10000]

    # Text / Code / Markdown fallback
    try:
        return file_bytes.decode("utf-8", errors="replace")[:30000]
    except Exception:
        return "[Fichier binaire non lisible]"


@router.post("", response_model=schemas.UserFileResponse, status_code=status.HTTP_201_CREATED)
async def upload_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    file_bytes = await file.read()
    if len(file_bytes) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"Fichier trop volumineux. Taille maximale : {MAX_FILE_SIZE // 1024 // 1024} Mo."
        )

    mime_type = file.content_type or "application/octet-stream"
    user_dir = UPLOADS_DIR / str(current_user.id)
    user_dir.mkdir(parents=True, exist_ok=True)
    safe_name = f"{uuid.uuid4().hex}_{Path(file.filename).name}"
    file_path = user_dir / safe_name
    file_path.write_bytes(file_bytes)

    parsed = parse_file_content(file_bytes, mime_type, file.filename)

    db_file = models.UserFile(
        user_id=current_user.id,
        filename=safe_name,
        original_name=file.filename,
        mime_type=mime_type,
        size_bytes=len(file_bytes),
        file_path=str(file_path),
        parsed_content=parsed
    )
    db.add(db_file)
    db.commit()
    db.refresh(db_file)
    return db_file


@router.get("", response_model=List[schemas.UserFileResponse])
def list_files(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    return (
        db.query(models.UserFile)
        .filter(models.UserFile.user_id == current_user.id)
        .order_by(models.UserFile.uploaded_at.desc())
        .all()
    )


@router.delete("/{file_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_file(
    file_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    db_file = db.query(models.UserFile).filter(
        models.UserFile.id == file_id,
        models.UserFile.user_id == current_user.id
    ).first()
    if not db_file:
        raise HTTPException(status_code=404, detail="Fichier non trouve.")
    try:
        Path(db_file.file_path).unlink(missing_ok=True)
    except Exception:
        pass
    db.delete(db_file)
    db.commit()
    return None
